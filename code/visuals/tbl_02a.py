import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

VISUAL_ID = "tbl_02a"
logger = logging.getLogger(__name__)


def _apply_formatting(df, params):
    if not params:
        return df

    df_formatted = df.copy()

    for key, fmt in params.items():
        if not key.startswith("format."):
            continue

        col = key.replace("format.", "")

        if col not in df_formatted.columns:
            continue

        if fmt == "int":
            df_formatted[col] = df_formatted[col].astype(int)

        elif fmt in ["comma", "float1", "float0", "percent"]:
            # ✅ keep numeric for Excel formatting
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

    return df_formatted


def _safe_numeric(series, fill_value=0):
    return pd.to_numeric(series, errors="coerce").fillna(fill_value)


def _calculate_mode(series):
    try:
        mode_series = series.mode(dropna=True)
        return mode_series.iloc[0] if not mode_series.empty else None
    except Exception:
        return None


def run(df, params, start_date, end_date, output_dir, generate_output_name):

    logger.info(f"[{VISUAL_ID}] Starting execution")

    # --------------------------------------------------
    # ✅ Required Columns
    # --------------------------------------------------

    required_cols = {
        "wheels_in_dtm",
        "wheels_out_dtm",
        "or_type"
    }

    missing = required_cols - set(df.columns)
    if missing:
        logger.warning(f"[{VISUAL_ID}] Missing required columns: {missing}")
        return

    if df.empty:
        logger.warning(f"[{VISUAL_ID}] Input dataframe is empty")
        return

    working_df = df.copy()

    logger.info(f"[{VISUAL_ID}] Input rows: {len(working_df):,}")

    # --------------------------------------------------
    # ✅ Datetime handling
    # --------------------------------------------------

    working_df["wheels_in_dtm"] = pd.to_datetime(
        working_df["wheels_in_dtm"], errors="coerce"
    )
    working_df["wheels_out_dtm"] = pd.to_datetime(
        working_df["wheels_out_dtm"], errors="coerce"
    )

    working_df = working_df.dropna(subset=["wheels_in_dtm", "wheels_out_dtm"])

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid datetime rows")
        return

    # --------------------------------------------------
    # ✅ Duration
    # --------------------------------------------------

    working_df["duration_minutes"] = (
        working_df["wheels_out_dtm"] - working_df["wheels_in_dtm"]
    ).dt.total_seconds() / 60.0

    working_df["duration_minutes"] = _safe_numeric(
        working_df["duration_minutes"]
    )

    working_df = working_df[working_df["duration_minutes"] > 0]

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid duration rows")
        return

    # --------------------------------------------------
    # ✅ Year
    # --------------------------------------------------

    working_df["year"] = working_df["wheels_out_dtm"].dt.year
    working_df = working_df.dropna(subset=["year"])

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid year values")
        return

    # --------------------------------------------------
    # ✅ Aggregation
    # --------------------------------------------------

    try:
        agg_df = (
            working_df
            .groupby(["year", "or_type"])
            .agg(
                encounter_count=("duration_minutes", "count"),
                min_duration=("duration_minutes", "min"),
                median_duration=("duration_minutes", "median"),
                mean_duration=("duration_minutes", "mean"),
                max_duration=("duration_minutes", "max"),
            )
            .reset_index()
        )

        mode_df = (
            working_df
            .groupby(["year", "or_type"])["duration_minutes"]
            .apply(_calculate_mode)
            .reset_index(name="mode_duration")
        )

        agg_df = agg_df.merge(mode_df, on=["year", "or_type"], how="left")

    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Aggregation failed: {e}")
        return

    if agg_df.empty:
        logger.warning(f"[{VISUAL_ID}] Aggregation returned empty dataset")
        return

    # --------------------------------------------------
    # ✅ Sorting
    # --------------------------------------------------

    agg_df = agg_df.sort_values(
        by=["year", "encounter_count"],
        ascending=[False, False]
    )

    for col in agg_df.columns:
        if col not in ["year", "or_type"]:
            agg_df[col] = pd.to_numeric(agg_df[col], errors="coerce")

    agg_df["year"] = agg_df["year"].astype(int)

    logger.info(f"[{VISUAL_ID}] Output rows: {len(agg_df):,}")

    # --------------------------------------------------
    # ✅ Output (Excel)
    # --------------------------------------------------

    try:
        filename = generate_output_name(
            visual_id=VISUAL_ID,
            start_date=start_date,
            end_date=end_date,
            cohort_id=(params or {}).get("cohort_id"),
            ext="xlsx"
        )

        output_path = os.path.join(output_dir, filename)

        output_df = _apply_formatting(agg_df, params)
        params = params or {}

        title = params.get("title", "")
        subtitle = params.get("cohort_desc")
        subtitle1 = params.get("subtitle1", "")

        # -----------------------
        # ✅ Write Excel
        # -----------------------
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            output_df.to_excel(
                writer,
                index=False,
                sheet_name="Sheet1",
                startrow=4
            )

        # -----------------------
        # ✅ Modify workbook
        # -----------------------
        wb = load_workbook(output_path)
        ws = wb.active

        # ✅ Titles
        ws["A1"] = title
        ws["A2"] = subtitle if subtitle else ""
        ws["A3"] = subtitle1 if subtitle1 else ""

        # ✅ Column widths
        for idx, col in enumerate(output_df.columns, start=1):
            col_letter = get_column_letter(idx)

            width = params.get(f"colwidth.{col}")
            if width:
                try:
                    ws.column_dimensions[col_letter].width = float(width)
                except Exception:
                    logger.warning(f"[{VISUAL_ID}] Invalid width for {col}")

        # ✅ Excel number formatting
        for idx, col in enumerate(output_df.columns, start=1):
            fmt = params.get(f"format.{col}")
            col_letter = get_column_letter(idx)

            if not fmt:
                continue

            for cell in ws[col_letter][1:]:  # skip header
                if cell.value is None:
                    continue

                if fmt == "comma":
                    cell.number_format = '#,##0'

                elif fmt == "float1":
                    cell.number_format = '#,##0.0'

                elif fmt == "float0":
                    cell.number_format = '#,##0'

                elif fmt == "percent":
                    cell.number_format = '0.0%'

        # ✅ Freeze header row
        ws.freeze_panes = "A6"

        wb.save(output_path)

        logger.info(f"[{VISUAL_ID}] Saved output to: {output_path}")

    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Failed to save output: {e}")
        return