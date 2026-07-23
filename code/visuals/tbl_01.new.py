# =============================================================================
# Domain      : Inpatient
# Report Name : Annual Patients and Patient Days Summary
#
# Description :
# Produces an annual inpatient utilization summary for the reporting year.
# Returns:
#   - Excel workbook
#   - RDB metrics
# =============================================================================

import os
import logging
import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

VISUAL_ID = "tbl_01"

logger = logging.getLogger(__name__)


def _safe_numeric(series, fill_value=0):
    return pd.to_numeric(
        series,
        errors="coerce"
    ).fillna(fill_value)


def _normalize_params(params):

    params = params or {}

    return {
        k: (
            None if pd.isna(v)
            else v
        )
        for k, v in params.items()
    }


def _apply_excel_formats(
    worksheet,
    dataframe,
    params
):

    for idx, column_name in enumerate(
        dataframe.columns,
        start=1
    ):

        fmt = params.get(
            f"format.{column_name}"
        )

        if not fmt:
            continue

        column_letter = get_column_letter(idx)

        for cell in worksheet[column_letter][5:]:

            if cell.value is None:
                continue

            if fmt == "comma":
                cell.number_format = "#,##0"

            elif fmt == "float1":
                cell.number_format = "#,##0.0"

            elif fmt == "float0":
                cell.number_format = "#,##0"

            elif fmt == "percent":
                cell.number_format = "0.0%"


def run(
    df,
    params,
    start_date,
    end_date,
    output_dir,
    generate_output_name
):

    logger.info(
        f"[{VISUAL_ID}] Starting execution"
    )

    params = _normalize_params(params)

    required_cols = {
        "discharge_fiscal_year",
        "count_of_patients",
        "total_days"
    }

    missing_cols = required_cols - set(df.columns)

    if missing_cols:

        logger.error(
            f"[{VISUAL_ID}] Missing required columns: "
            f"{sorted(missing_cols)}"
        )

        return

    if df.empty:

        logger.warning(
            f"[{VISUAL_ID}] Input dataframe is empty"
        )

        return

    working_df = df.copy()

    logger.info(
        f"[{VISUAL_ID}] Input rows: "
        f"{len(working_df):,}"
    )

    working_df["count_of_patients"] = _safe_numeric(
        working_df["count_of_patients"]
    )

    working_df["total_days"] = _safe_numeric(
        working_df["total_days"]
    )

    working_df["discharge_fiscal_year"] = pd.to_numeric(
        working_df["discharge_fiscal_year"],
        errors="coerce"
    )

    working_df = working_df.dropna(
        subset=["discharge_fiscal_year"]
    )

    if working_df.empty:

        logger.warning(
            f"[{VISUAL_ID}] No valid rows after cleaning"
        )

        return

    report_year = pd.to_datetime(
        end_date,
        errors="coerce"
    ).year

    year_type = str(
        params.get("year_type", "")
    ).strip().lower()

    if year_type == "fiscal":

        working_df = working_df[
            working_df["discharge_fiscal_year"]
            == report_year
        ]

        logger.info(
            f"[{VISUAL_ID}] Applied fiscal-year filter "
            f"for {report_year}"
        )

    else:

        logger.warning(
            f"[{VISUAL_ID}] Unsupported year_type: "
            f"{year_type}"
        )

    if working_df.empty:

        logger.warning(
            f"[{VISUAL_ID}] No rows after year filter"
        )

        return

    working_df["year"] = report_year

    agg_df = (
        working_df
        .groupby(
            "year",
            as_index=False
        )
        .agg(
            patients=(
                "count_of_patients",
                "sum"
            ),
            patient_days=(
                "total_days",
                "sum"
            )
        )
    )

    if agg_df.empty:

        logger.warning(
            f"[{VISUAL_ID}] Aggregation returned no rows"
        )

        return

    agg_df["year"] = (
        pd.to_numeric(
            agg_df["year"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    agg_df["patients"] = (
        pd.to_numeric(
            agg_df["patients"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    agg_df["patient_days"] = (
        pd.to_numeric(
            agg_df["patient_days"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    filename = generate_output_name(
        visual_id=VISUAL_ID,
        start_date=start_date,
        end_date=end_date,
        cohort_id=params.get(
            "cohort_id"
        ),
        ext="xlsx"
    )

    output_path = os.path.join(
        output_dir,
        filename
    )

    title = params.get(
        "title",
        "Annual Patients and Patient Days Summary"
    )

    subtitle = (
        params.get("cohort_desc")
        or ""
    )

    subtitle1 = params.get(
        "subtitle1",
        ""
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        agg_df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=4,
            index=False
        )

    wb = load_workbook(output_path)
    ws = wb.active

    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A3"] = subtitle1

    for idx, column_name in enumerate(
        agg_df.columns,
        start=1
    ):

        width = params.get(
            f"colwidth.{column_name}"
        )

        if width is None:
            continue

        try:

            ws.column_dimensions[
                get_column_letter(idx)
            ].width = float(width)

        except Exception:

            logger.warning(
                f"[{VISUAL_ID}] Invalid column width "
                f"for {column_name}"
            )

    _apply_excel_formats(
        ws,
        agg_df,
        params
    )

    ws.freeze_panes = "A6"

    wb.save(output_path)

    logger.info(
        f"[{VISUAL_ID}] Output written: "
        f"{output_path}"
    )

    year_prefix = {
        "fiscal": "FY",
        "calendar": "CY"
    }.get(year_type, "")

    rdb_rows = []

    for _, row in agg_df.iterrows():

        rdb_rows.append({

            "run_id":
                params.get("run_id"),

            "visual_id":
                VISUAL_ID,

            "client_name":
                params.get("client_name"),

            "domain":
                params.get("domain"),

            "cohort_id":
                params.get("cohort_id"),

            "domain_cohort":
                f"{params.get('domain')}.{params.get('cohort_id')}",

            "dimension":
                "year",

            "dimension_value":
                int(row["year"]),

            "dimension_value_label":
                f"{year_prefix}{int(row['year'])}",

            "secondary_dimension":
                "",

            "secondary_dimension_value":
                "",

            "metric":
                "patients",

            "metric_type":
                "count",

            "value":
                int(row["patients"]),

            "start_date":
                start_date,

            "end_date":
                end_date,

            "report_title":
                title
        })

        rdb_rows.append({

            "run_id":
                params.get("run_id"),

            "visual_id":
                VISUAL_ID,

            "client_name":
                params.get("client_name"),

            "domain":
                params.get("domain"),

            "cohort_id":
                params.get("cohort_id"),

            "domain_cohort":
                f"{params.get('domain')}.{params.get('cohort_id')}",

            "dimension":
                "year",

            "dimension_value":
                int(row["year"]),

            "dimension_value_label":
                f"{year_prefix}{int(row['year'])}",

            "secondary_dimension":
                "",

            "secondary_dimension_value":
                "",

            "metric":
                "patient_days",

            "metric_type":
                "count",

            "value":
                int(row["patient_days"]),

            "start_date":
                start_date,

            "end_date":
                end_date,

            "report_title":
                title
        })

    return {
        "output_path": output_path,
        "rdb": rdb_rows
    }