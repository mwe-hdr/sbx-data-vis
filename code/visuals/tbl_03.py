import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

VISUAL_ID = "tbl_03"
logger = logging.getLogger(__name__)


def _apply_formatting(df, params):
    if not params:
        return df

    df_formatted = df.copy()

    for key, fmt in params.items():
        if not key.startswith("format."):
            continue

        col = key.replace("format.", "")

        if col not in df_formatted.columns:
            continue

        if fmt == "int":
            df_formatted[col] = df_formatted[col].astype(int)

        elif fmt == "comma":
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

        elif fmt == "float1":
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

        elif fmt == "float0-":
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

        elif fmt == "percent":
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

    return df_formatted


def _safe_numeric(series, fill_value=0):
    return pd.to_numeric(series, errors="coerce").fillna(fill_value)


def _calculate_mode(series):
    """
    Return first mode value (handles multimodal results safely)
    """
    try:
        mode_series = series.mode(dropna=True)
        return mode_series.iloc[0] if not mode_series.empty else None
    except Exception:
        return None


def run(df, params, start_date, end_date, output_dir, generate_output_name):
    """
    Table 03: ED Encounter Duration Summary by Year

    Spec:
    - Group by year derived from ed_stop_dtm
    - duration_minutes = wheels_out_dtm - ed_start_dtm
    - stats:
        min, median, mode, mean, max
    """

    logger.info(f"[{VISUAL_ID}] Starting execution")

    # --------------------------------------------------
    # ✅ Required Columns
    # --------------------------------------------------

    logger.info(f"[{VISUAL_ID}] Columns:")
    logger.info(list(df.columns))

    required_cols = {
        "ed_start_dtm",
        "ed_stop_dtm"
    }

    missing = required_cols - set(df.columns)
    if missing:
        logger.warning(f"[{VISUAL_ID}] Missing required columns: {missing}")
        return

    if df.empty:
        logger.warning(f"[{VISUAL_ID}] Input dataframe is empty")
        return

    working_df = df.copy()

    # --------------------------------------------------
    # ✅ Date Conversion
    # --------------------------------------------------

    working_df["ed_start_dtm"] = pd.to_datetime(
        working_df["ed_start_dtm"], errors="coerce"
    )

    working_df["ed_stop_dtm"] = pd.to_datetime(
        working_df["ed_stop_dtm"], errors="coerce"
    )

    working_df = working_df.dropna(
        subset=["ed_start_dtm", "ed_stop_dtm"]
    )

    # --------------------------------------------------
    # ✅ Apply Date Filter (from driver)
    # Include encounters whose ED time window overlaps
    # the reporting period.
    # --------------------------------------------------

    start_dt = (
        pd.to_datetime(start_date, errors="coerce")
        if start_date else None
    )

    end_dt = (
        pd.to_datetime(end_date, errors="coerce")
        if end_date else None
    )

    if start_dt is not None and end_dt is not None:

        working_df = working_df[
            (working_df["ed_start_dtm"] <= end_dt) &
            (working_df["ed_stop_dtm"] >= start_dt)
        ]

    elif start_dt is not None:

        working_df = working_df[
            working_df["ed_stop_dtm"] >= start_dt
        ]

    elif end_dt is not None:

        working_df = working_df[
            working_df["ed_start_dtm"] <= end_dt
        ]

    logger.info(
        f"[{VISUAL_ID}] After overlap filter: "
        f"{len(working_df):,} rows "
        f"(start={start_date}, end={end_date})"
    )

    logger.info(f"[{VISUAL_ID}] Input rows: {len(working_df):,}")

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid datetime rows")
        return

    # --------------------------------------------------
    # ✅ Duration Calculation (minutes)
    # --------------------------------------------------

    working_df["duration_minutes"] = (
        working_df["ed_stop_dtm"] - working_df["ed_start_dtm"]
    ).dt.total_seconds() / 60.0

    working_df["duration_minutes"] = _safe_numeric(
        working_df["duration_minutes"], fill_value=0
    )

    # Remove negatives or zero durations
    working_df = working_df[working_df["duration_minutes"] > 0]

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid duration rows")
        return

    # --------------------------------------------------
    # ✅ Report Year
    # --------------------------------------------------

    report_year = pd.to_datetime(
        end_date,
        errors="coerce"
    ).year

    working_df["year"] = report_year

    logger.info(
        f"[{VISUAL_ID}] Assigned report_year={report_year}"
    )

    # --------------------------------------------------
    # ✅ Aggregation
    # --------------------------------------------------

    try:
        agg_df = (
            working_df
            .groupby("year")
            .agg(
                encounter_count=("duration_minutes", "count"), 
                min_duration=("duration_minutes", "min"),
                median_duration=("duration_minutes", "median"),
                mean_duration=("duration_minutes", "mean"),
                max_duration=("duration_minutes", "max"),
            )
            .reset_index()
        )

        # Mode requires separate handling
        mode_df = (
            working_df
            .groupby("year")["duration_minutes"]
            .apply(_calculate_mode)
            .reset_index(name="mode_duration")
        )

        agg_df = agg_df.merge(mode_df, on="year", how="left")

    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Aggregation failed: {e}")
        return

    if agg_df.empty:
        logger.warning(f"[{VISUAL_ID}] Aggregation returned empty dataset")
        return

    # --------------------------------------------------
    # ✅ Sorting
    # --------------------------------------------------

    agg_df = agg_df.sort_values(
        by="year",
        ascending=False
    )

    # Ensure numeric consistency
    for col in agg_df.columns:
        if col != "year":
            agg_df[col] = pd.to_numeric(agg_df[col], errors="coerce")

    agg_df["year"] = agg_df["year"].astype(int)

    logger.info(f"[{VISUAL_ID}] Output rows: {len(agg_df):,}")

    # --------------------------------------------------
    # ✅ Output
    # --------------------------------------------------

    try:
        filename = generate_output_name(
            visual_id=VISUAL_ID,
            start_date=start_date,
            end_date=end_date,
            cohort_id=(params or {}).get("cohort_id"),
            ext="xlsx"
        )

        output_path = os.path.join(output_dir, filename)

        output_df = _apply_formatting(agg_df, params)
        params = params or {}

        title = params.get("title", "")
        subtitle = params.get("cohort_desc")
        subtitle1 = params.get("subtitle1", "")
        client_name = params.get("client_name")
        year_type = params.get("year_type")
        year_prefix = {
            "fiscal": "FY",
            "calendar": "CY"
        }.get(str(params.get("year_type", "")).lower(), "")

        # -----------------------
        # Write Excel
        # -----------------------
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            output_df.to_excel(
                writer,
                index=False,
                sheet_name="Sheet1",
                startrow=4
            )

        # -----------------------
        # Modify workbook
        # -----------------------
        wb = load_workbook(output_path)
        ws = wb.active

        # ✅ Title / subtitle rows
        ws["A1"] = title
        ws["A2"] = subtitle if subtitle else ""
        ws["A3"] = subtitle1 if subtitle1 else ""

        # ✅ Apply column widths
        for idx, col in enumerate(output_df.columns, start=1):
            col_letter = get_column_letter(idx)

            width = params.get(f"colwidth.{col}")
            if width:
                try:
                    ws.column_dimensions[col_letter].width = float(width)
                except Exception:
                    logger.warning(f"[{VISUAL_ID}] Invalid width for {col}")

        # ✅ Apply formats
        for idx, col in enumerate(output_df.columns, start=1):
            fmt = params.get(f"format.{col}")
            col_letter = get_column_letter(idx)

            if not fmt:
                continue

            for cell in ws[col_letter][1:]:  

                if cell.value is None:
                    continue

                if fmt == "comma":
                    cell.number_format = '#,##0'

                elif fmt == "float1":
                    cell.number_format = '#,##0.0'

                elif fmt == "float0":
                    cell.number_format = '#,##0'                

                elif fmt == "percent":
                    cell.number_format = '0.0%'        

        # ✅ Freeze header row (row 5 in Excel)
        ws.freeze_panes = "A6"

        wb.save(output_path)

        logger.info(f"[{VISUAL_ID}] Saved output to: {output_path}")

        rdb_rows = []

        for _, row in agg_df.iterrows():
            base_record = {
                "run_id": params.get("run_id"),
                "visual_id": VISUAL_ID,
                "client_name": params.get("client_name"),
                "domain": params.get("domain"),
                "cohort_id": params.get("cohort_id"),
                "domain_cohort":
                    f"{params.get('domain')}.{params.get('cohort_id')}",
                "dimension": "year",
                "dimension_value": row["year"],
                "dimension_value_label":
                    f"{year_prefix}{int(row['year'])}",
                "secondary_dimension": "",
                "secondary_dimension_value": "",
                "start_date": start_date,
                "end_date": end_date,
                "report_title": params.get("title")
            }
            rdb_rows.append({
                **base_record,
                "metric": "encounter_count",
                "metric_type": "count",
                "value": row["encounter_count"]
            })
            rdb_rows.append({
                **base_record,
                "metric": "min_duration",
                "metric_type": "duration_minutes",
                "value": row["min_duration"]
            })
            rdb_rows.append({
                **base_record,
                "metric": "median_duration",
                "metric_type": "duration_minutes",
                "value": row["median_duration"]
            })
            rdb_rows.append({
                **base_record,
                "metric": "mean_duration",
                "metric_type": "duration_minutes",
                "value": row["mean_duration"]
            })
            rdb_rows.append({
                **base_record,
                "metric": "mode_duration",
                "metric_type": "duration_minutes",
                "value": row["mode_duration"]
            })
            rdb_rows.append({
                **base_record,
                "metric": "max_duration",
                "metric_type": "duration_minutes",
                "value": row["max_duration"]
            })


    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Failed to save output: {e}")
        return
    
    return {
        "output_path": output_path,
        "rdb": rdb_rows
    }