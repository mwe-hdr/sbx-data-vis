# =============================================================================
# Domain      : Inpatient
# Report Name : Annual Patients and Patient Days Summary
#
# Description :
# Produces an annual inpatient utilization summary for the reporting year.
# The program filters data to the selected fiscal year and aggregates total
# patient volume and total patient days across the population being
# reported. Results are written to a formatted Excel workbook and returned
# as reporting database (RDB) records for downstream analytics.
#
# This report provides a high-level view of inpatient activity and census
# utilization by summarizing:
#   - Total patients discharged
#   - Total inpatient days
#
# The output is commonly used for executive reporting, utilization review,
# volume trending, and benchmarking across reporting periods or cohorts.
#
# Inputs :
#   - discharge_fiscal_year : Fiscal year of patient discharge
#   - count_of_patients     : Patient count measure
#   - total_days            : Total inpatient days
#   - start_date            : Reporting period start date
#   - end_date              : Reporting period end date
#
# Outputs :
#   - Excel table summarizing:
#       * Reporting year
#       * Total patients
#       * Total patient days
#   - RDB records containing annual patient and patient day metrics
# =============================================================================

import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

VISUAL_ID = "tbl_01"
logger = logging.getLogger(__name__)


def _apply_formatting(df, params):
    if not params:
        return df

    df_formatted = df.copy()

    for key, fmt in params.items():
        if not key.startswith("format."):
            continue

        col = key.replace("format.", "")

        if col not in df_formatted.columns:
            continue

        if fmt == "int":
            df_formatted[col] = df_formatted[col].astype(int)

        elif fmt in ["comma", "float1", "float0", "percent"]:
            # ✅ keep numeric (no string formatting)
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

    return df_formatted


def _safe_numeric(series, fill_value=0):
    return pd.to_numeric(series, errors="coerce").fillna(fill_value)


def run(df, params, start_date, end_date, output_dir, generate_output_name):

    logger.info(f"[{VISUAL_ID}] Starting execution")

    params = params or {}
    params = {
        k: (None if pd.isna(v) else v)
        for k, v in params.items()
    }

    # --------------------------------------------------
    # ✅ Required Columns
    # --------------------------------------------------

    required_cols = {
        "discharge_fiscal_year",
        "count_of_patients",
        "total_days"
    }

    missing = required_cols - set(df.columns)
    if missing:
        logger.warning(f"[{VISUAL_ID}] Missing required columns: {missing}")
        return

    if df.empty:
        logger.warning(f"[{VISUAL_ID}] Input dataframe is empty")
        return

    working_df = df.copy()

    logger.info(f"[{VISUAL_ID}] Input rows: {len(working_df):,}")

    # --------------------------------------------------
    # ✅ Data Preparation
    # --------------------------------------------------

    working_df["count_of_patients"] = _safe_numeric(
        working_df["count_of_patients"], fill_value=0
    )

    working_df["total_days"] = _safe_numeric(
        working_df["total_days"], fill_value=0
    )

    working_df["discharge_fiscal_year"] = pd.to_numeric(
        working_df["discharge_fiscal_year"], errors="coerce"
    )

    working_df = working_df.dropna(subset=["discharge_fiscal_year"])

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid rows after cleaning")
        return

    # --------------------------------------------------
    # ✅ Apply Date Filter (from driver)
    # --------------------------------------------------

    report_year = pd.to_datetime(
        end_date,
        errors="coerce"
    ).year

    year_type = str(
        params.get("year_type", "")
    ).lower()

    if year_type == "fiscal":

        working_df = working_df[
            working_df["discharge_fiscal_year"] == report_year
        ]

        logger.info(
            f"[{VISUAL_ID}] Fiscal year filter applied: "
            f"discharge_fiscal_year == {report_year}"
        )

    else:

        logger.warning(
            f"[{VISUAL_ID}] Unsupported year_type: {year_type}"
        )

    working_df["year"] = report_year

    # --------------------------------------------------
    # ✅ Aggregation
    # --------------------------------------------------

    try:
        agg_df = (
            working_df
            .groupby("year", as_index=False)
            .agg(
                patients=("count_of_patients", "sum"),
                patient_days=("total_days", "sum")
            )
        )
    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Aggregation failed: {e}")
        return

    if agg_df.empty:
        logger.warning(f"[{VISUAL_ID}] Aggregation returned empty dataset")
        return

    # --------------------------------------------------
    # ✅ Sorting
    # --------------------------------------------------

    agg_df = agg_df.sort_values(
        by="year",
        ascending=False
    )

    for col in ["year", "patients", "patient_days"]:
        agg_df[col] = pd.to_numeric(agg_df[col], errors="coerce").fillna(0)

    agg_df["year"] = agg_df["year"].astype(int)
    agg_df["patients"] = agg_df["patients"].astype(int)
    agg_df["patient_days"] = agg_df["patient_days"].astype(int)

    logger.info(f"[{VISUAL_ID}] Output rows: {len(agg_df):,}")

    # --------------------------------------------------
    # ✅ Output (Excel)
    # --------------------------------------------------

    try:
        filename = generate_output_name(
            visual_id=VISUAL_ID,
            start_date=start_date,
            end_date=end_date,
            cohort_id=(params or {}).get("cohort_id"),
            ext="xlsx"
        )

        output_path = os.path.join(output_dir, filename)

        output_df = _apply_formatting(agg_df, params)
        params = params or {}

        title = params.get("title", "")
        subtitle = params.get("cohort_desc")
        subtitle1 = params.get("subtitle1", "")
        client_name = params.get("client_name")
        year_type = params.get("year_type")
        year_prefix = {
            "fiscal": "FY",
            "calendar": "CY"
        }.get(str(params.get("year_type", "")).lower(), "")

        # -----------------------
        # ✅ Write Excel
        # -----------------------
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            output_df.to_excel(
                writer,
                index=False,
                sheet_name="Sheet1",
                startrow=4
            )

        # -----------------------
        # ✅ Modify workbook
        # -----------------------
        wb = load_workbook(output_path)
        ws = wb.active

        # ✅ Titles
        ws["A1"] = title
        ws["A2"] = subtitle if subtitle else ""
        ws["A3"] = subtitle1 if subtitle1 else ""

        # ✅ Column widths
        for idx, col in enumerate(output_df.columns, start=1):
            col_letter = get_column_letter(idx)

            width = params.get(f"colwidth.{col}")
            if width:
                try:
                    ws.column_dimensions[col_letter].width = float(width)
                except Exception:
                    logger.warning(f"[{VISUAL_ID}] Invalid width for {col}")

        # ✅ Excel number formatting
        for idx, col in enumerate(output_df.columns, start=1):
            fmt = params.get(f"format.{col}")
            col_letter = get_column_letter(idx)

            if not fmt:
                continue

            for cell in ws[col_letter][1:]:  # skip header
                if cell.value is None:
                    continue

                if fmt == "comma":
                    cell.number_format = '#,##0'

                elif fmt == "float1":
                    cell.number_format = '#,##0.0'

                elif fmt == "float0":
                    cell.number_format = '#,##0'

                elif fmt == "percent":
                    cell.number_format = '0.0%'

        # ✅ Freeze pane
        ws.freeze_panes = "A6"

        wb.save(output_path)

        logger.info(f"[{VISUAL_ID}] Saved output to: {output_path}")

    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Failed to save output: {e}")
        return
    
    rdb_rows = []

    for _, row in agg_df.iterrows():

        rdb_rows.append({
            "run_id": params.get("run_id"),
            "visual_id": VISUAL_ID,
            "client_name": params.get("client_name"),

            "domain": params.get("domain"),
            "cohort_id": params.get("cohort_id"),

            "domain_cohort":
                f"{params.get('domain')}.{params.get('cohort_id')}",

            "dimension": "year",
            "dimension_value": row["year"],
            "dimension_value_label":
                f"{year_prefix}{int(row['year'])}",

            "secondary_dimension": "",
            "secondary_dimension_value": "",

            "metric": "patients",
            "metric_type": "count",
            "value": row["patients"],

            "start_date": start_date,
            "end_date": end_date,

            "report_title": params.get("title")
        })

        rdb_rows.append({
            "run_id": params.get("run_id"),
            "visual_id": VISUAL_ID,
            "client_name": params.get("client_name"),

            "domain": params.get("domain"),
            "cohort_id": params.get("cohort_id"),

            "domain_cohort":
                f"{params.get('domain')}.{params.get('cohort_id')}",

            "dimension": "year",
            "dimension_value": row["year"],
            "dimension_value_label":
                f"{year_prefix}{int(row['year'])}",

            "secondary_dimension": "",
            "secondary_dimension_value": "",

            "metric": "patient_days",
            "metric_type": "count",
            "value": row["patient_days"],

            "start_date": start_date,
            "end_date": end_date,

            "report_title": params.get("title")
        })

    return {
    "output_path": output_path,
    "rdb": rdb_rows
    }