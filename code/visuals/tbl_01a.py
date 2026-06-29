import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

VISUAL_ID = "tbl_01a"
logger = logging.getLogger(__name__)


def _apply_formatting(df, params):
    if not params:
        return df

    df_formatted = df.copy()

    for key, fmt in params.items():
        if not key.startswith("format."):
            continue

        col = key.replace("format.", "")

        if col not in df_formatted.columns:
            continue

        if fmt == "int":
            df_formatted[col] = df_formatted[col].astype(int)

        elif fmt in ["comma", "float1", "float0", "percent"]:
            df_formatted[col] = pd.to_numeric(df_formatted[col], errors="coerce")

    return df_formatted


def _safe_numeric(series, fill_value=0):
    return pd.to_numeric(series, errors="coerce").fillna(fill_value)


def run(df, params, start_date, end_date, output_dir, generate_output_name):

    logger.info(f"[{VISUAL_ID}] Starting execution")

    # --------------------------------------------------
    # ✅ Required Columns
    # --------------------------------------------------

    required_cols = {
        "discharge_fiscal_year",
        "count_of_patients",
        "total_days",
        "service_line"
    }

    missing = required_cols - set(df.columns)
    if missing:
        logger.warning(f"[{VISUAL_ID}] Missing required columns: {missing}")
        return

    if df.empty:
        logger.warning(f"[{VISUAL_ID}] Input dataframe is empty")
        return

    working_df = df.copy()

    logger.info(f"[{VISUAL_ID}] Input rows: {len(working_df):,}")

    # --------------------------------------------------
    # ✅ Data Preparation
    # --------------------------------------------------

    working_df["count_of_patients"] = _safe_numeric(
        working_df["count_of_patients"], fill_value=0
    )

    working_df["total_days"] = _safe_numeric(
        working_df["total_days"], fill_value=0
    )

    working_df["discharge_fiscal_year"] = pd.to_numeric(
        working_df["discharge_fiscal_year"], errors="coerce"
    )

    working_df = working_df.dropna(subset=["discharge_fiscal_year", "service_line"])

    if working_df.empty:
        logger.warning(f"[{VISUAL_ID}] No valid rows after cleaning")
        return

    # --------------------------------------------------
    # ✅ Aggregation
    # --------------------------------------------------

    try:
        agg_df = (
            working_df
            .groupby(["discharge_fiscal_year", "service_line"], as_index=False)
            .agg(
                patients=("count_of_patients", "sum"),
                patient_days=("total_days", "sum")
            )
        )
    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Aggregation failed: {e}")
        return

    if agg_df.empty:
        logger.warning(f"[{VISUAL_ID}] Aggregation returned empty dataset")
        return

    # --------------------------------------------------
    # ✅ Sorting
    # --------------------------------------------------

    agg_df = agg_df.sort_values(
        by=["discharge_fiscal_year", "patients"],
        ascending=[False, False]
    )

    for col in ["discharge_fiscal_year", "patients", "patient_days"]:
        agg_df[col] = pd.to_numeric(agg_df[col], errors="coerce").fillna(0)

    agg_df["discharge_fiscal_year"] = agg_df["discharge_fiscal_year"].astype(int)
    agg_df["patients"] = agg_df["patients"].astype(int)
    agg_df["patient_days"] = agg_df["patient_days"].astype(int)

    logger.info(f"[{VISUAL_ID}] Output rows: {len(agg_df):,}")

    # --------------------------------------------------
    # ✅ Output (Excel)
    # --------------------------------------------------

    try:
        filename = generate_output_name(
            visual_id=VISUAL_ID,
            start_date=start_date,
            end_date=end_date,
            cohort_id=(params or {}).get("cohort_id"),
            ext="xlsx"
        )

        output_path = os.path.join(output_dir, filename)

        output_df = _apply_formatting(agg_df, params)
        params = params or {}

        title = params.get("title", "")
        subtitle = params.get("cohort_desc")
        subtitle1 = params.get("subtitle1", "")

        # -----------------------
        # ✅ Write Excel
        # -----------------------
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            output_df.to_excel(
                writer,
                index=False,
                sheet_name="Sheet1",
                startrow=4
            )

        # -----------------------
        # ✅ Modify workbook
        # -----------------------
        wb = load_workbook(output_path)
        ws = wb.active

        # ✅ Titles
        ws["A1"] = title
        ws["A2"] = subtitle if subtitle else ""
        ws["A3"] = subtitle1 if subtitle1 else ""

        # ✅ Column widths
        for idx, col in enumerate(output_df.columns, start=1):
            col_letter = get_column_letter(idx)

            width = params.get(f"colwidth.{col}")
            if width:
                try:
                    ws.column_dimensions[col_letter].width = float(width)
                except Exception:
                    logger.warning(f"[{VISUAL_ID}] Invalid width for {col}")

        # ✅ Excel number formatting
        for idx, col in enumerate(output_df.columns, start=1):
            fmt = params.get(f"format.{col}")
            col_letter = get_column_letter(idx)

            if not fmt:
                continue

            for cell in ws[col_letter][1:]:  # skip header row
                if cell.value is None:
                    continue

                if fmt == "comma":
                    cell.number_format = '#,##0'

                elif fmt == "float1":
                    cell.number_format = '#,##0.0'

                elif fmt == "float0":
                    cell.number_format = '#,##0'

                elif fmt == "percent":
                    cell.number_format = '0.0%'

        # ✅ Freeze pane
        ws.freeze_panes = "A6"

        wb.save(output_path)

        logger.info(f"[{VISUAL_ID}] Saved output to: {output_path}")

    except Exception as e:
        logger.error(f"[{VISUAL_ID}] Failed to save output: {e}")
        return